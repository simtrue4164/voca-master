import openpyxl
import json
import os

file_path = "MVP Vol. 1 어휘 목록★.xlsx"

if not os.path.exists(file_path):
    print(f"Error: {file_path} not found.")
    exit(1)

wb = openpyxl.load_workbook(file_path, data_only=True)
all_words_by_day = {}

# User says English words from B3.
# Need to identify how "DAY" is separated. 
# Usually, they are in different sheets or have a "DAY" column.
# Let's assume there are sheets named "DAY 1", "DAY 2", etc. or a single sheet with a separator.

for sheet_name in wb.sheetnames:
    sheet = wb[sheet_name]
    words = []
    # Starting from B3 (row 3, col 2)
    for row in range(3, sheet.max_row + 1):
        word = sheet.cell(row=row, column=2).value
        if word and isinstance(word, str) and word.strip():
            words.append(word.strip())
        if len(words) >= 50:
            break
    
    if words:
        all_words_by_day[sheet_name] = words

# Write to a JSON for verification
with open("words_summary.json", "w", encoding="utf-8") as f:
    json.dump(all_words_by_day, f, ensure_ascii=False, indent=2)

print(f"Extraction complete. {len(all_words_by_day)} days found.")
